# app/ata_window.py

from PyQt5.QtWidgets import QDialog, QPushButton, QVBoxLayout, QLabel, QHBoxLayout, QScrollArea, QWidget, QComboBox
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from utils import resource_path
from database_operations import get_connection

class DataWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ver Datos")
        # Eliminamos setFixedSize para permitir redimensionar

        self.day_button = QPushButton("Tiempo por Día")
        self.day_button.clicked.connect(self.show_day_data)
        self.month_button = QPushButton("Tiempo por Mes")
        self.month_button.clicked.connect(self.show_month_data)
        self.year_button = QPushButton("Tiempo por Año")
        self.year_button.clicked.connect(self.show_year_data)
        self.data_label = QLabel()
        self.return_button = QPushButton("Regresar")
        self.return_button.clicked.connect(self.close)
        
        self.week_combo = QComboBox()
        self.week_combo.addItems([str(i) for i in range(1, 53)])  # Weeks 1-52
        self.week_combo.setCurrentText("52")  # Default to week 52
        
        # Añadir QComboBox para el año
        self.year_combo = QComboBox()
        self.year_combo.addItems([str(i) for i in range(2023, 2026)])  # Añadir años desde 2023 hasta 2025 (ajusta según sea necesario)
        self.year_combo.setCurrentText(str(datetime.now().year))  # Establecer el año actual como predeterminado
        
        self.plot_button = QPushButton("Mostrar Gráfico")
        self.plot_button.clicked.connect(self.show_plot)
        

        layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        button_layout.addWidget(self.day_button)
        button_layout.addWidget(self.month_button)
        button_layout.addWidget(self.year_button)
        layout.addLayout(button_layout)
        
        plot_layout = QHBoxLayout()
        plot_layout.addWidget(QLabel("Año:"))
        plot_layout.addWidget(self.year_combo)
        plot_layout.addWidget(QLabel("Semana:"))
        plot_layout.addWidget(self.week_combo)
        plot_layout.addWidget(self.plot_button)
        layout.addLayout(plot_layout)
        
        self.figure = Figure(figsize=(10, 6))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.addWidget(self.data_label)
        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)

        layout.addWidget(self.return_button)
        self.setLayout(layout)

        self.excel_file = resource_path("basedatos_pro.xlsx")
        self.sheet_name = "Hoja1"


    def show_plot(self):
        selected_week = int(self.week_combo.currentText())
        # Obtener el año seleccionado
        selected_year = int(self.year_combo.currentText())
        
        conn = get_connection()
        cursor = conn.cursor()

        try:
            query = """
            SELECT PROYECTO, TAREA, TIEMPO_INICIAL, TIEMPO_FINAL 
            FROM Registro_Tiempo
            """
            cursor.execute(query)
            rows = cursor.fetchall()
            
            # Crear DataFrame manualmente
            df = pd.DataFrame(rows, columns=['PROYECTO', 'TAREA', 'TIEMPO_INICIAL', 'TIEMPO_FINAL'])
            print("DataFrame original (df):")
            print(df)
            
            df['TIEMPO INICIAL'] = pd.to_datetime(df['TIEMPO_INICIAL'], format='%Y-%m-%d %H:%M:%S')
            df['TIEMPO FINAL'] = pd.to_datetime(df['TIEMPO_FINAL'], format='%Y-%m-%d %H:%M:%S')

            df['DURACION'] = (df['TIEMPO FINAL'] - df['TIEMPO INICIAL']).dt.total_seconds() / 3600
            # Filtrar por año antes de calcular la semana
            df['AÑO'] = df['TIEMPO INICIAL'].dt.year
            df = df[df['AÑO'] == selected_year]
            
            df['SEMANA'] = df['TIEMPO INICIAL'].dt.isocalendar().week
            df['DIA SEMANA'] = df['TIEMPO INICIAL'].dt.dayofweek

            df_week = df[df['SEMANA'] == selected_week]
            print(f"DataFrame filtrado por año {selected_year} y semana {selected_week} (df_week):")
            print(df_week)
            
            df_week_days = df_week.groupby(['DIA SEMANA', 'TAREA'])['DURACION'].sum().unstack()
            df_week_days = df_week_days.fillna(0)

            dias_semana = {0:'Lunes', 1:'Martes', 2:'Miércoles', 3:'Jueves', 4:'Viernes', 5:'Sábado', 6:'Domingo'}
            df_week_days = df_week_days.rename(index=dias_semana)
            df_week_days = df_week_days.reindex(dias_semana.values())

            self.figure.clear()
            ax = self.figure.add_subplot(111)
            df_week_days.plot(kind='bar', stacked=True, ax=ax)
            ax.set_title(f'Horas dedicadas a cada proyecto por día de la semana (Semana {selected_week})')
            ax.set_xlabel('Día de la Semana')
            ax.set_ylabel('Horas')

            max_hours = df_week_days.sum(axis=1).max()
            ax.set_ylim(0, max(7, max_hours))

            total_hours = 17  # Horas totales de trabajo por día

            for rect in ax.patches:
                height = rect.get_height()
                width = rect.get_width()
                x = rect.get_x()
                y = rect.get_y()
                label_x = x + width / 2
                label_y = y + height / 2

                if height > 0:
                    percentage = (height / total_hours) * 100
                    label = f'{percentage:.1f}%\n({height:.1f} hrs)'
                    ax.text(label_x, label_y, label, ha='center', va='center', color='white', fontsize=8, fontweight='bold')

            total_perc = (df_week_days.sum(axis=1)/total_hours)*100
            for idx, value in enumerate(total_perc):
                ax.text(idx, df_week_days.sum(axis=1)[idx] + 0.1, f'{value:.1f}%\n({df_week_days.sum(axis=1)[idx]:.1f} hrs)', 
                        ha='center', va='bottom', fontsize=8, fontweight='bold')

            plt.tight_layout()
            self.canvas.draw()
        except Exception as e:
            print(f"Error al leer datos de la base de datos: {e}")
        finally:
            cursor.close()
            conn.close()

    def show_day_data(self):
        workbook = load_workbook(self.excel_file)
        sheet = workbook[self.sheet_name]

        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            project = row[1]
            task = row[2]
            start_time = datetime.strptime(row[3], "%Y-%m-%d %H:%M:%S")
            end_time = datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
            duration = end_time - start_time

            date = start_time.strftime("%Y-%m-%d")
            if date not in data:
                data[date] = {}
            if project not in data[date]:
                data[date][project] = {}
            if task not in data[date][project]:
                data[date][project][task] = duration.total_seconds()
            else:
                data[date][project][task] += duration.total_seconds()

        text = ""
        for date, projects in data.items():
            text += f"Fecha: {date}\n"
            for project, tasks in projects.items():
                text += f"  Proyecto: {project}\n"
                for task, seconds in tasks.items():
                    hours = seconds / 3600
                    text += f"    Tarea: {task} - Tiempo: {hours:.2f} horas\n"
            text += "\n"

        self.data_label.setText(text)

    def show_month_data(self):
        workbook = load_workbook(self.excel_file)
        sheet = workbook[self.sheet_name]

        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            project = row[1]
            task = row[2]
            start_time = datetime.strptime(row[3], "%Y-%m-%d %H:%M:%S")
            end_time = datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
            duration = end_time - start_time

            month = start_time.strftime("%Y-%m")
            if month not in data:
                data[month] = {}
            if project not in data[month]:
                data[month][project] = {}
            if task not in data[month][project]:
                data[month][project][task] = duration.total_seconds()
            else:
                data[month][project][task] += duration.total_seconds()

        text = ""
        for month, projects in data.items():
            text += f"Mes: {month}\n"
            for project, tasks in projects.items():
                text += f"  Proyecto: {project}\n"
                for task, seconds in tasks.items():
                    hours = seconds / 3600
                    text += f"    Tarea: {task} - Tiempo: {hours:.2f} horas\n"
            text += "\n"

        self.data_label.setText(text)

    def show_year_data(self):
        workbook = load_workbook(self.excel_file)
        sheet = workbook[self.sheet_name]

        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            project = row[1]
            task = row[2]
            start_time = datetime.strptime(row[3], "%Y-%m-%d %H:%M:%S")
            end_time = datetime.strptime(row[4], "%Y-%m-%d %H:%M:%S")
            duration = end_time - start_time

            year = start_time.strftime("%Y")
            if year not in data:
                data[year] = {}
            if project not in data[year]:
                data[year][project] = {}
            if task not in data[year][project]:
                data[year][project][task] = duration.total_seconds()
            else:
                data[year][project][task] += duration.total_seconds()

        text = ""
        for year, projects in data.items():
            text += f"Año: {year}\n"
            for project, tasks in projects.items():
                text += f"  Proyecto: {project}\n"
                for task, seconds in tasks.items():
                    hours = seconds / 3600
                    text += f"    Tarea: {task} - Tiempo: {hours:.2f} horas\n"
            text += "\n"

        self.data_label.setText(text)